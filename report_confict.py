from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from itertools import combinations
from dataclasses import dataclass
from pathlib import Path
from typing import Any, DefaultDict, Dict, Iterable, List, Sequence

from openpyxl import Workbook, load_workbook

PV_KEY_PATTERN = re.compile(r"^pv_(\d+)\.(.+)$", re.IGNORECASE)
COMMENT_PATTERN = re.compile(r"^#\s*(\d+)\s*-\s*(.+)$")
WHITESPACE_PATTERN = re.compile(r"\s+")


def normalize_title(raw: str) -> str:
    return WHITESPACE_PATTERN.sub(" ", raw).strip().lower()


@dataclass
class SongEntry:
    pv_id: int
    title: str
    title_en: str | None
    source_name: str
    source_type: str
    pvdb_path: Path | None = None

    @property
    def normalized_title(self) -> str:
        reference = self.title_en or ""
        return normalize_title(reference)

    @property
    def source_label(self) -> str:
        return f"{self.source_type}:{self.source_name}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Scan pv_db files inside mod folders, detect PV ID and song title conflicts, "
            "and export the results to Excel."
        )
    )
    parser.add_argument(
        "--mods",
        required=True,
        type=Path,
        help="Path to the directory that contains mod folders.",
    )
    parser.add_argument(
        "--base-catalog",
        default=Path("data/data.xlsx"),
        type=Path,
        help="Excel file that lists the base game's PV IDs and titles.",
    )
    parser.add_argument(
        "--output",
        default=Path("reports/conflict_report.xlsx"),
        type=Path,
        help="Where to write the generated Excel report.",
    )
    return parser.parse_args()


def load_base_catalog(catalog_path: Path) -> List[SongEntry]:
    if not catalog_path.exists():
        print(f"[warn] Base catalog {catalog_path} not found, skipping.")
        return []

    workbook = load_workbook(catalog_path, data_only=True)
    worksheet = workbook.active

    header_row_index = None
    headers: Sequence[str] | None = None
    for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if any(cell is not None and str(cell).strip() for cell in row):
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            header_row_index = idx
            break

    if headers is None or header_row_index is None:
        print(f"[warn] Base catalog {catalog_path} is empty.")
        return []

    id_idx = _find_column(headers, ("pv_id", "id"))
    name_idx = _find_column(headers, ("song_name", "title", "name"))
    name_en_idx = _find_column(headers, ("song_name_en", "title_en", "name_en"))

    if id_idx is None or name_idx is None:
        print(
            f"[warn] Could not find both ID and name columns in {catalog_path}. "
            "Expected headers containing 'id' and 'name'."
        )
        return []

    entries: List[SongEntry] = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        raw_id = row[id_idx] if id_idx < len(row) else None
        if raw_id is None:
            continue
        try:
            pv_id = int(str(raw_id).strip().lstrip("#"))
        except ValueError:
            continue

        primary_name = _safe_string(row, name_idx)
        secondary_name = _safe_string(row, name_en_idx) if name_en_idx is not None else None
        title = primary_name or secondary_name
        if not title:
            continue

        entries.append(
            SongEntry(
                pv_id=pv_id,
                title=title,
                title_en=secondary_name,
                source_name="base-game",
                source_type="base",
            )
        )

    workbook.close()
    return entries


def _safe_string(row: Sequence[object], idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    value = row[idx]
    if value is None:
        return None
    return str(value).strip()


def _find_column(headers: Sequence[str], candidates: Sequence[str]) -> int | None:
    for candidate in candidates:
        for idx, header in enumerate(headers):
            if candidate in header:
                return idx
    return None


def discover_pvdb_files(mods_root: Path) -> List[tuple[str, Path]]:
    files: List[tuple[str, Path]] = []
    for mod_dir in mods_root.iterdir():
        if not mod_dir.is_dir():
            continue
        mod_name = mod_dir.name
        for path in mod_dir.rglob("*.txt"):
            if not path.is_file():
                continue
            if path.parent.name != "rom":
                continue
            lowered = path.name.lower()
            if not lowered == "mod_pv_db.txt":
                continue
            files.append((mod_name, path))
    return files


def parse_pvdb_file(pvdb_path: Path, mod_name: str) -> List[SongEntry]:
    song_data: Dict[int, Dict[str, str]] = defaultdict(dict)
    with pvdb_path.open("r", encoding="utf-8", errors="ignore") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            match = COMMENT_PATTERN.match(line)
            if match:
                pv_id = int(match.group(1))
                song_data[pv_id]["comment_title"] = match.group(2).strip()
                continue
            if line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            key_match = PV_KEY_PATTERN.match(key)
            if not key_match:
                continue
            pv_id = int(key_match.group(1))
            attr = key_match.group(2).lower()
            if attr == "song_name":
                song_data[pv_id]["song_name"] = value
            elif attr == "song_name_en":
                song_data[pv_id]["song_name_en"] = value

    entries: List[SongEntry] = []
    for pv_id, data in song_data.items():
        primary = data.get("song_name") or data.get("comment_title")
        secondary = data.get("song_name_en")
        title = primary or secondary
        if not title:
            continue
        entries.append(
            SongEntry(
                pv_id=pv_id,
                title=title,
                title_en=secondary,
                source_name=mod_name,
                source_type="mod",
                pvdb_path=pvdb_path,
            )
        )
    return entries


def detect_id_conflicts(entries: Iterable[SongEntry]) -> Dict[int, List[SongEntry]]:
    grouped: Dict[int, List[SongEntry]] = defaultdict(list)
    for entry in entries:
        grouped[entry.pv_id].append(entry)

    conflicts: Dict[int, List[SongEntry]] = {}
    for pv_id, group in grouped.items():
        sources = {(item.source_type, item.source_name) for item in group}
        if len(sources) > 1:
            conflicts[pv_id] = group
    return conflicts


def detect_song_conflicts(entries: Iterable[SongEntry]) -> Dict[str, List[SongEntry]]:
    grouped: Dict[str, List[SongEntry]] = defaultdict(list)
    for entry in entries:
        key = entry.normalized_title
        if not key:
            continue
        grouped[key].append(entry)

    conflicts: Dict[str, List[SongEntry]] = {}
    for normalized_title, group in grouped.items():
        sources = {(item.source_type, item.source_name) for item in group}
        if len(sources) > 1:
            conflicts[normalized_title] = group
    return conflicts


def _format_song_names(group: Sequence[SongEntry]) -> str:
    names = [entry.title_en or entry.title for entry in group if entry.title_en or entry.title]
    if not names:
        return ""
    return ", ".join(sorted(set(names)))


def _format_sources(group: Sequence[SongEntry]) -> str:
    return ", ".join(sorted({entry.source_label for entry in group}))


def _build_conflict_rows(
    id_conflicts: Dict[int, List[SongEntry]],
    song_conflicts: Dict[str, List[SongEntry]],
) -> List[List[str]]:
    rows: List[List[str]] = []
    for pv_id, group in sorted(id_conflicts.items()):
        rows.append(
            [
                "id_conflict",
                _format_song_names(group),
                str(pv_id),
                _format_sources(group),
            ]
        )

    for normalized_title, group in sorted(song_conflicts.items()):
        pv_ids = ", ".join(str(pid) for pid in sorted({entry.pv_id for entry in group}))
        display_name = _format_song_names(group) or normalized_title
        rows.append(
            [
                "song_conflict",
                display_name,
                pv_ids,
                _format_sources(group),
            ]
        )
    return rows


def _build_pack_conflict_rows(
    entries: Sequence[SongEntry],
    id_conflicts: Dict[int, List[SongEntry]],
    song_conflicts: Dict[str, List[SongEntry]],
) -> List[List[Any]]:
    songs_by_SP: Dict[str, List[SongEntry]] = defaultdict(list)
    for entry in entries:
        songs_by_SP[entry.source_name].append(entry)

    def _stat_record():
        return {
            "total": 0,
            "conflict_keys": set(),
            "partners": defaultdict(set),
        }

    stats: DefaultDict[str, Dict[str, Any]] = defaultdict(_stat_record)
    for sp_name, songs in songs_by_SP.items():
        stats[sp_name]["total"] = len(songs)

    def register_conflict(group: Sequence[SongEntry], key: tuple[str, object]) -> None:
        source_map: Dict[str, List[SongEntry]] = defaultdict(list)
        for item in group:
            source_map[item.source_name].append(item)
        if len(source_map) < 2:
            return
        labels = sorted(source_map.keys())
        for label in labels:
            stats[label]["conflict_keys"].add(key)
        for left, right in combinations(labels, 2):
            stats[left]["partners"][right].add(key)
            stats[right]["partners"][left].add(key)

    for pv_id, group in id_conflicts.items():
        register_conflict(group, ("id", pv_id))

    for normalized_title, group in song_conflicts.items():
        register_conflict(group, ("song", normalized_title))

    rows: List[List[Any]] = []
    for pack_name in sorted(songs_by_SP.keys()):
        record = stats[pack_name]
        total = record["total"]
        conflict_total = len(record["conflict_keys"])
        rows.append([pack_name, conflict_total, "", total])
        partner_map = record["partners"]
        for partner_name, partner_conflict_keys in sorted(partner_map.items()):
            rows.append([pack_name, len(partner_conflict_keys), partner_name, songs_by_SP[partner_name].__len__()])
    return rows


def export_report(
    output_path: Path,
    entries: List[SongEntry],
    id_conflicts: Dict[int, List[SongEntry]],
    song_conflicts: Dict[str, List[SongEntry]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    songs_sheet = workbook.active
    songs_sheet.title = "songs"
    songs_sheet.append([
        "pv_id",
        "title",
        "title_en",
        "source_type",
        "source_name",
        "pvdb_path",
    ])
    for entry in sorted(entries, key=lambda item: (item.source_type, item.source_name, item.pv_id)):
        songs_sheet.append(
            [
                entry.pv_id,
                entry.title,
                entry.title_en or "",
                entry.source_type,
                entry.source_name,
                str(entry.pvdb_path) if entry.pvdb_path else "",
            ]
        )

    conflicts_sheet = workbook.create_sheet("conflicts")
    conflicts_sheet.append(["conflict_type", "song_name", "pv_ids", "sources"])
    for row in _build_conflict_rows(id_conflicts, song_conflicts):
        conflicts_sheet.append(row)

    pack_sheet = workbook.create_sheet("pack_conflicts")
    pack_sheet.append(["pack_name", "conflict_count", "conflict_partner", "total_songs"])
    for row in _build_pack_conflict_rows(entries, id_conflicts, song_conflicts):
        pack_sheet.append(row)

    workbook.save(output_path)
    workbook.close()


def main() -> None:
    args = parse_args()
    mods_root = args.mods.expanduser().resolve()
    if not mods_root.exists():
        raise SystemExit(f"Mods path {mods_root} does not exist.")

    # base_catalog = args.base_catalog.expanduser()
    output_path = args.output.expanduser()

    # base_entries = load_base_catalog(base_catalog)
    mod_entries: List[SongEntry] = []
    pvdb_files = discover_pvdb_files(mods_root)
    if not pvdb_files:
        print("[warn] No pv_db files found under the provided mods directory.")
    else:
        print(f"[info] Found {len(pvdb_files)} pv_db files in mods directory.")
    for mod_name, pvdb_path in pvdb_files:
        parsed = parse_pvdb_file(pvdb_path, mod_name)
        if not parsed:
            continue
        mod_entries.extend(parsed)

    # all_entries = [*base_entries, *mod_entries]
    all_entries = mod_entries
    if not mod_entries:
        print("[warn] No mod songs detected. Report will include base catalog only.")

    mod_counter = Counter(entry.source_name for entry in mod_entries)
    if mod_counter:
        print("[info] Songs per mod:")
        for mod_name, count in mod_counter.most_common():
            print(f"  - {mod_name}: {count} songs")

    print(f"[info] Total songs indexed: {len(all_entries)}")

    id_conflicts = detect_id_conflicts(all_entries)
    song_conflicts = detect_song_conflicts(all_entries)

    if id_conflicts:
        print("[conflict] PV ID clashes detected:")
        for pv_id, group in sorted(id_conflicts.items()):
            details = "; ".join(f"{entry.title} ({entry.source_label})" for entry in group)
            print(f"  - {pv_id}: {details}")
    else:
        print("[ok] No PV ID conflicts found.")

    if song_conflicts:
        print("[conflict] Song title clashes detected:")
        for normalized_title, group in sorted(song_conflicts.items()):
            human_title = group[0].title
            details = "; ".join(f"id {entry.pv_id} ({entry.source_label})" for entry in group)
            print(f"  - {human_title}: {details}")
    else:
        print("[ok] No song title conflicts found.")

    export_report(output_path, all_entries, id_conflicts, song_conflicts)
    print(f"[info] Report saved to {output_path}")


if __name__ == "__main__":
    main()
