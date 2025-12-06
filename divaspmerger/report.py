from __future__ import annotations

from itertools import combinations
from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook

from .logging_utils import log_conflict, log_ok
from .models import ConflictRecord, ConflictType, PackInfo, ResolutionPlan, SongEntry


def print_conflict_details(
    id_conflicts: Dict[int, List[SongEntry]], song_conflicts: Dict[str, List[SongEntry]]
) -> None:
    if id_conflicts:
        log_conflict("PV ID clashes detected:")
        for pv_id, group in sorted(id_conflicts.items()):
            details = "; ".join(
                f"{entry.title} ({entry.source_label})" for entry in group
            )
            log_conflict(f"{pv_id}: {details}", indent=2)
    else:
        log_ok("No PV ID conflicts found.")
    if song_conflicts:
        log_conflict("Song title clashes detected:")
        for normalized_title, group in sorted(song_conflicts.items()):
            human_title = group[0].title
            details = "; ".join(
                f"id {entry.pv_id} ({entry.source_label})" for entry in group
            )
            log_conflict(f"{human_title}: {details}", indent=2)
    else:
        log_ok("No song title conflicts found.")


def _format_song_names(group: Sequence[SongEntry]) -> str:
    names = [entry.title_en or entry.title for entry in group if entry.title_en or entry.title]
    if not names:
        return ""
    return ", ".join(sorted(set(names)))


def _format_sources(group: Sequence[SongEntry]) -> str:
    return ", ".join(sorted({entry.source_label for entry in group}))


def _build_conflict_rows(
    conflicts: List[ConflictRecord],
) -> List[List[str]]:
    rows: List[List[str]] = []
    for conflict in conflicts:
        if conflict.conflict_type == ConflictType.ID:
            pv_ids = str(conflict.winner.pv_id)
            song_names = _format_song_names(conflict.entries)
        elif conflict.conflict_type == ConflictType.SONG:
            pv_ids = ", ".join(str(entry.pv_id) for entry in conflict.entries)
            song_names = conflict.winner.title_en or conflict.winner.title
        else:
            pv_ids = ", ".join(str(entry.pv_id) for entry in conflict.entries)
            song_names = _format_song_names(conflict.entries)
        row = [
            conflict.conflict_type.value,  # conflict_type
            song_names,                     # song_names
            pv_ids,                         # pv_ids
            conflict.winner.source_label,  # picked_source
            _format_sources(conflict.entries),  # sources
        ]
        rows.append(row)
    return rows


def _build_pack_conflict_rows(
    conflicts: List[ConflictRecord],
    pack_infos: Dict[str, PackInfo],
    plans: Dict[str, ResolutionPlan],
) -> List[List[Any]]:
    
    conflict_partners: Dict[str, set[str]] = {}
    conflict_ids: Dict[str, set[int]] = {}
    for conflict in conflicts:
        involved_mods = conflict.involved_mods()
        for mod_a, mod_b in combinations(involved_mods, 2):
            conflict_partners.setdefault(mod_a, set()).add(mod_b)
            conflict_partners.setdefault(mod_b, set()).add(mod_a)
        for mod in involved_mods:
            conflict_ids.setdefault(mod, set()).add(conflict.winner.pv_id)
    
    rows: List[List[Any]] = []
    for pack_name, pack_info in pack_infos.items():
        plan = plans.get(pack_name)
        conflict_partner = conflict_partners.get(pack_name, [])
        removal_songs = plan.total_removals if plan else 0
        row = [
            pack_info.priority,  # priority
            pack_name,  # pack_name
            pack_info.num_songs,  # total_songs
            len(conflict_ids.get(pack_name, [])),  # conflict_songs
            removal_songs,  # removal_songs
            pack_info.num_songs - removal_songs,  # remain_songs
            ", ".join(conflict_partner),  # conflict_partners
        ]
        rows.append(row)
    return sorted(rows, key=lambda r: (r[0], r[1]))  # Sort by priority, then pack_name


def export_report(
    output_path: Path,
    entries: List[SongEntry],
    pack_infos: Dict[str, PackInfo],
    conflicts: List[ConflictRecord],
    plans: Dict[str, ResolutionPlan],
) -> None:
    """Write an Excel report summarizing detected conflicts."""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    
    # Export Song info sheet
    songs_sheet = workbook.active
    if not songs_sheet:
        songs_sheet = workbook.create_sheet("songs")
    else:
        songs_sheet.title = "songs"
    songs_sheet.append([
        "pv id",
        "title",
        "title en",
        "source type",
        "source name",
        "pvdb path",
    ])
    for entry in sorted(entries, key=lambda item: (item.source_type, item.source_name, item.pv_id)):
        songs_sheet.append(
            [
                entry.pv_id,
                entry.title,
                entry.title_en or "",
                entry.source_type,
                entry.source_name,
                str(entry.pvdb_path) if entry.pvdb_path else "",
            ]
        )

    # Export Conflicts sheet
    conflicts_sheet = workbook.create_sheet("conflicts")
    conflicts_sheet.append(["conflict type", "song name", "pv ids", "picked source", "all sources"])
    for row in _build_conflict_rows(conflicts=conflicts):
        conflicts_sheet.append(row)

    # Export Pack Conflicts sheet
    pack_sheet = workbook.create_sheet("pack_conflicts")
    pack_sheet.append(
        [
            "priority",
            "pack name",
            "total songs",
            "conflict songs",
            "removal songs",
            "remain songs",
            "conflict partners",
        ]
    )
    for row in _build_pack_conflict_rows(conflicts=conflicts, pack_infos=pack_infos, plans=plans):
        pack_sheet.append(row)

    workbook.save(output_path)
    workbook.close()


__all__ = ["print_conflict_details", "export_report"]
